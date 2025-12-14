from __future__ import annotations
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable
import json

from openpyxl import load_workbook, Workbook

@dataclass
class InputRow:
    product_code: str
    numerical_order: float
    slug: str

def read_input_rows(xlsx_bytes: bytes) -> list[InputRow]:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # ищем колонки по заголовкам
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if not val:
            continue
        headers[str(val).strip().lower()] = col

    if "product_code" not in headers or "numerical_order" not in headers or "slug" not in headers:
        raise RuntimeError("Input Excel must contain headers: product_code, numerical_order, slug")

    out: list[InputRow] = []
    for row in range(2, ws.max_row + 1):
        pc = ws.cell(row=row, column=headers["product_code"]).value
        no = ws.cell(row=row, column=headers["numerical_order"]).value
        slug = ws.cell(row=row, column=headers["slug"]).value
        if pc is None or no is None or slug is None:
            continue

        pc = str(pc).strip()
        slug_s = str(slug).strip().strip("/")
        if not pc or not slug_s:
            continue

        try:
            no_f = float(no)
        except Exception:
            continue

        out.append(InputRow(product_code=pc, numerical_order=no_f, slug=slug_s))
    return out

def build_output_xlsx(rows: Iterable[tuple[str, float | None]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "prices"

    ws.append(["Код товара Tabletki.ua", "Цена"])
    for product_code, price in rows:
        ws.append([product_code, price])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_output_json(records: Iterable[dict]) -> bytes:
    """Build UTF-8 JSON output.

    `records` should be an iterable of dicts, e.g.:
      {"code": "1094251", "city": "kiev", "delivery_price": 923.0}

    Returns bytes suitable for saving as .json.
    """
    data = list(records)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    return payload.encode("utf-8")